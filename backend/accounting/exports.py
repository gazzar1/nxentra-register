"""
Export utilities for accounting data.
Supports Excel (.xlsx), CSV (.csv), and Text (.txt) formats.
"""
import csv
import io
from datetime import datetime
from decimal import Decimal
from typing import Any, Callable

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExportFormat:
    EXCEL = 'xlsx'
    CSV = 'csv'
    TXT = 'txt'

    CHOICES = [EXCEL, CSV, TXT]
    CONTENT_TYPES = {
        EXCEL: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        CSV: 'text/csv',
        TXT: 'text/plain',
    }


def format_value(value: Any) -> str:
    """Format a value for export."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    return str(value)


def export_to_excel(
    data: list[dict],
    columns: list[dict],
    title: str = 'Export',
    sheet_name: str = 'Data',
) -> bytes:
    """
    Export data to Excel format.

    Args:
        data: List of dictionaries containing the data
        columns: List of column definitions with 'key', 'header', and optional 'width'
        title: Title for the export (used in header row)
        sheet_name: Name of the worksheet

    Returns:
        Bytes of the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    # Export timestamp
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    timestamp_cell = ws.cell(row=2, column=1, value=f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    timestamp_cell.alignment = Alignment(horizontal='center')
    timestamp_cell.font = Font(italic=True, size=10, color='666666')

    # Header row
    header_row = 4
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col['header'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        # Set column width
        width = col.get('width', 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, row_data in enumerate(data, header_row + 1):
        for col_idx, col in enumerate(columns, 1):
            value = row_data.get(col['key'], '')
            cell = ws.cell(row=row_idx, column=col_idx, value=format_value(value))
            cell.border = thin_border

            # Right-align numeric columns
            if col.get('numeric'):
                cell.alignment = Alignment(horizontal='right')

    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_to_csv(
    data: list[dict],
    columns: list[dict],
    delimiter: str = ',',
) -> str:
    """
    Export data to CSV format.

    Args:
        data: List of dictionaries containing the data
        columns: List of column definitions with 'key' and 'header'
        delimiter: CSV delimiter character

    Returns:
        CSV string
    """
    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)

    # Header row
    writer.writerow([col['header'] for col in columns])

    # Data rows
    for row_data in data:
        row = [format_value(row_data.get(col['key'], '')) for col in columns]
        writer.writerow(row)

    return output.getvalue()


def export_to_txt(
    data: list[dict],
    columns: list[dict],
    separator: str = '\t',
) -> str:
    """
    Export data to text format with fixed-width or tab-separated columns.

    Args:
        data: List of dictionaries containing the data
        columns: List of column definitions with 'key', 'header', and optional 'width'
        separator: Column separator character

    Returns:
        Text string
    """
    lines = []

    # Calculate column widths for fixed-width format
    col_widths = []
    for col in columns:
        width = col.get('width', len(col['header']))
        # Check actual data widths
        for row_data in data:
            value = format_value(row_data.get(col['key'], ''))
            width = max(width, len(value))
        col_widths.append(min(width, 50))  # Cap at 50 chars

    # Header row
    header_parts = []
    for idx, col in enumerate(columns):
        header_parts.append(col['header'].ljust(col_widths[idx]))
    lines.append(separator.join(header_parts))

    # Separator line
    sep_parts = ['-' * col_widths[idx] for idx in range(len(columns))]
    lines.append(separator.join(sep_parts))

    # Data rows
    for row_data in data:
        row_parts = []
        for idx, col in enumerate(columns):
            value = format_value(row_data.get(col['key'], ''))
            # Truncate if too long
            if len(value) > col_widths[idx]:
                value = value[:col_widths[idx] - 3] + '...'
            row_parts.append(value.ljust(col_widths[idx]))
        lines.append(separator.join(row_parts))

    return '\n'.join(lines)


def create_export_response(
    data: list[dict],
    columns: list[dict],
    format: str,
    filename: str,
    title: str = 'Export',
) -> HttpResponse:
    """
    Create an HTTP response with the exported file.

    Args:
        data: List of dictionaries containing the data
        columns: List of column definitions
        format: Export format (xlsx, csv, txt)
        filename: Base filename (without extension)
        title: Title for Excel export

    Returns:
        HttpResponse with the file content
    """
    if format not in ExportFormat.CHOICES:
        raise ValueError(f"Invalid format: {format}. Must be one of {ExportFormat.CHOICES}")

    content_type = ExportFormat.CONTENT_TYPES[format]
    full_filename = f"{filename}.{format}"

    if format == ExportFormat.EXCEL:
        content = export_to_excel(data, columns, title=title)
        response = HttpResponse(content, content_type=content_type)
    elif format == ExportFormat.CSV:
        content = export_to_csv(data, columns)
        response = HttpResponse(content, content_type=content_type)
        response.charset = 'utf-8-sig'  # BOM for Excel compatibility
    else:  # TXT
        content = export_to_txt(data, columns)
        response = HttpResponse(content, content_type=content_type)

    response['Content-Disposition'] = f'attachment; filename="{full_filename}"'
    return response


# =============================================================================
# Chart of Accounts Export Configuration
# =============================================================================

ACCOUNT_EXPORT_COLUMNS = [
    {'key': 'code', 'header': 'Account Code', 'width': 15},
    {'key': 'name', 'header': 'Account Name', 'width': 30},
    {'key': 'name_ar', 'header': 'Account Name (Arabic)', 'width': 30},
    {'key': 'account_type', 'header': 'Account Type', 'width': 15},
    {'key': 'normal_balance', 'header': 'Normal Balance', 'width': 12},
    {'key': 'status', 'header': 'Status', 'width': 10},
    {'key': 'parent_code', 'header': 'Parent Code', 'width': 15},
    {'key': 'is_header', 'header': 'Is Header', 'width': 10},
    {'key': 'description', 'header': 'Description', 'width': 40},
    {'key': 'balance', 'header': 'Balance', 'width': 15, 'numeric': True},
]

ACCOUNT_EXPORT_COLUMNS_SIMPLE = [
    {'key': 'code', 'header': 'Account Code', 'width': 15},
    {'key': 'name', 'header': 'Account Name', 'width': 30},
    {'key': 'account_type', 'header': 'Account Type', 'width': 15},
    {'key': 'normal_balance', 'header': 'Normal Balance', 'width': 12},
    {'key': 'status', 'header': 'Status', 'width': 10},
    {'key': 'balance', 'header': 'Balance', 'width': 15, 'numeric': True},
]


def prepare_account_export_data(accounts, include_balance: bool = True) -> list[dict]:
    """Prepare account data for export."""
    data = []
    for account in accounts:
        row = {
            'code': account.code,
            'name': account.name,
            'name_ar': account.name_ar or '',
            'account_type': account.account_type,
            'normal_balance': account.normal_balance,
            'status': account.status,
            'parent_code': account.parent.code if account.parent else '',
            'is_header': account.is_header,
            'description': account.description or '',
        }
        if include_balance:
            try:
                row['balance'] = account.accountbalance.balance if hasattr(account, 'accountbalance') else Decimal('0.00')
            except Exception:
                row['balance'] = Decimal('0.00')
        data.append(row)
    return data


# =============================================================================
# Journal Entries Export Configuration
# =============================================================================

JOURNAL_ENTRY_EXPORT_COLUMNS = [
    {'key': 'entry_number', 'header': 'Entry Number', 'width': 15},
    {'key': 'date', 'header': 'Date', 'width': 12},
    {'key': 'period', 'header': 'Period', 'width': 8},
    {'key': 'memo', 'header': 'Memo', 'width': 35},
    {'key': 'status', 'header': 'Status', 'width': 12},
    {'key': 'kind', 'header': 'Kind', 'width': 12},
    {'key': 'total_debit', 'header': 'Total Debit', 'width': 15, 'numeric': True},
    {'key': 'total_credit', 'header': 'Total Credit', 'width': 15, 'numeric': True},
    {'key': 'currency', 'header': 'Currency', 'width': 8},
    {'key': 'source_module', 'header': 'Source Module', 'width': 15},
    {'key': 'source_document', 'header': 'Source Document', 'width': 20},
    {'key': 'created_by_email', 'header': 'Created By', 'width': 25},
    {'key': 'created_at', 'header': 'Created At', 'width': 18},
]

JOURNAL_LINE_EXPORT_COLUMNS = [
    {'key': 'entry_number', 'header': 'Entry Number', 'width': 15},
    {'key': 'entry_date', 'header': 'Entry Date', 'width': 12},
    {'key': 'line_no', 'header': 'Line #', 'width': 8},
    {'key': 'account_code', 'header': 'Account Code', 'width': 15},
    {'key': 'account_name', 'header': 'Account Name', 'width': 30},
    {'key': 'description', 'header': 'Description', 'width': 35},
    {'key': 'debit', 'header': 'Debit', 'width': 15, 'numeric': True},
    {'key': 'credit', 'header': 'Credit', 'width': 15, 'numeric': True},
    {'key': 'currency', 'header': 'Currency', 'width': 8},
    {'key': 'memo', 'header': 'Entry Memo', 'width': 30},
]


def prepare_journal_entry_export_data(entries) -> list[dict]:
    """Prepare journal entry data for export (summary level)."""
    data = []
    for entry in entries:
        row = {
            'entry_number': entry.entry_number or entry.public_id,
            'date': entry.date,
            'period': entry.period,
            'memo': entry.memo or '',
            'status': entry.status,
            'kind': entry.kind,
            'total_debit': entry.total_debit,
            'total_credit': entry.total_credit,
            'currency': entry.currency,
            'source_module': entry.source_module or '',
            'source_document': entry.source_document or '',
            'created_by_email': entry.created_by.email if entry.created_by else '',
            'created_at': entry.created_at,
        }
        data.append(row)
    return data


def prepare_journal_lines_export_data(entries) -> list[dict]:
    """Prepare journal entry lines data for export (detailed level)."""
    data = []
    for entry in entries:
        for line in entry.lines.all().select_related('account').order_by('line_no'):
            row = {
                'entry_number': entry.entry_number or entry.public_id,
                'entry_date': entry.date,
                'line_no': line.line_no,
                'account_code': line.account.code,
                'account_name': line.account.name,
                'description': line.description or '',
                'debit': line.debit,
                'credit': line.credit,
                'currency': line.currency or entry.currency,
                'memo': entry.memo or '',
            }
            data.append(row)
    return data
