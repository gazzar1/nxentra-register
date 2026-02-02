# edim/parsers.py
"""
File parsers for EDIM ingestion.

Parsers convert uploaded files (CSV, XLSX, JSON) into lists of dicts
representing raw records.
"""

import csv
import json
import io
from typing import Tuple, List, Dict, Any


def parse_csv(file) -> List[Dict[str, Any]]:
    """
    Parse a CSV file into a list of dicts.

    Args:
        file: File-like object containing CSV data

    Returns:
        List of dicts, one per row
    """
    # Handle both binary and text modes
    if hasattr(file, 'read'):
        content = file.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8-sig')  # Handle BOM
        file = io.StringIO(content)

    reader = csv.DictReader(file)
    records = []
    for row in reader:
        # Clean up keys (remove BOM, strip whitespace)
        cleaned_row = {
            k.strip().lstrip('\ufeff'): v.strip() if isinstance(v, str) else v
            for k, v in row.items()
            if k is not None
        }
        records.append(cleaned_row)
    return records


def parse_xlsx(file) -> List[Dict[str, Any]]:
    """
    Parse an Excel file into a list of dicts.

    Args:
        file: File-like object containing Excel data

    Returns:
        List of dicts, one per row
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel file parsing. Install with: pip install openpyxl")

    # Read file content
    if hasattr(file, 'read'):
        content = file.read()
        if hasattr(file, 'seek'):
            file.seek(0)
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    else:
        workbook = openpyxl.load_workbook(file, data_only=True)

    sheet = workbook.active
    if sheet is None:
        return []

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    # First row is header
    headers = [str(h).strip() if h else f"column_{i}" for i, h in enumerate(rows[0])]

    records = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # Skip empty rows
        record = {}
        for i, value in enumerate(row):
            if i < len(headers):
                # Convert Excel values to strings for consistency
                if value is None:
                    record[headers[i]] = ""
                elif isinstance(value, (int, float)):
                    # Preserve numeric precision
                    record[headers[i]] = str(value) if not isinstance(value, float) or value == int(value) else str(value)
                else:
                    record[headers[i]] = str(value).strip()
        records.append(record)

    return records


def parse_json(file) -> List[Dict[str, Any]]:
    """
    Parse a JSON file into a list of dicts.

    Args:
        file: File-like object containing JSON data

    Returns:
        List of dicts, one per record
    """
    if hasattr(file, 'read'):
        content = file.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8')
    else:
        content = file

    data = json.loads(content)

    # Handle both array and object with records key
    if isinstance(data, list):
        records = data
    elif isinstance(data, dict):
        # Look for common keys that contain the records
        for key in ['records', 'data', 'items', 'rows', 'entries']:
            if key in data and isinstance(data[key], list):
                records = data[key]
                break
        else:
            # Treat single object as single record
            records = [data]
    else:
        raise ValueError("JSON must be an array or object")

    # Ensure all records are dicts
    result = []
    for record in records:
        if isinstance(record, dict):
            result.append(record)
        else:
            raise ValueError("Each record in JSON must be an object")

    return result


def detect_and_parse(file, filename: str) -> Tuple[str, List[Dict[str, Any]]]:
    """
    Detect file type from extension and parse accordingly.

    Args:
        file: File-like object
        filename: Original filename (used for type detection)

    Returns:
        Tuple of (file_type, records)
    """
    extension = filename.lower().split(".")[-1] if "." in filename else ""

    if extension == "csv":
        return "csv", parse_csv(file)
    elif extension in ("xlsx", "xls"):
        return "xlsx", parse_xlsx(file)
    elif extension == "json":
        return "json", parse_json(file)
    else:
        raise ValueError(f"Unsupported file type: {extension}. Supported: csv, xlsx, xls, json")
